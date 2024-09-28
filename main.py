import openpyxl.workbook
import pandas as pd
import openpyxl
import os

# 240907C70PXA8X

class Main:
    def __init__(self) -> None:
        pass
   
    def consulta_preco(produto):
        """Produto = [
            0 - nome
            1 - variacao
            2 - sku
        ]

        """
        workbook_new = openpyxl.load_workbook("RELATORIO_VENDAS.xlsx")
        sheet_produtos = workbook_new["PRODUTOS"]
        df = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="PRODUTOS")
        
        linha = 0
        valor = 0
        
        for produtos in sheet_produtos['C']:
            
            if produto[0] == produtos.value and produto[1] == sheet_produtos['D'][linha].value:
                valor = sheet_produtos['E'][linha].value
                return float(valor)
            
            linha += 1        

        valor = Main.save_produtos([produto[2], # sku
                                    produto[0],  # nome 
                                    produto[1]]) # variacao
        return float(valor)
        
    
    def save_produtos(produto):
        """SALVA PRODUTOS NA TABELA RELATORIO VENDAS

        Args:
            produto (_list_): Eecebe lista com informações do produto
        """
        workbook_new = openpyxl.load_workbook("RELATORIO_VENDAS.xlsx")
        sheet_produtos = workbook_new["PRODUTOS"]
        df = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="PRODUTOS")
        nome = produto[1]
        variacao = produto[2]
        sku = produto[0]
        custo = 0
        
        linha_inserir = sheet_produtos.max_row + 1

        if variacao == "" or variacao == "nan":
            variacao = "S/V"
       
        sheet_produtos[f"B{linha_inserir}"] = f"{sku}"   #  SKU
        sheet_produtos[f"C{linha_inserir}"] = f"{nome}"   # Nome
        sheet_produtos[f"D{linha_inserir}"] = f"{variacao}"   # variação
        os.system("cls")
        print(f"\n -> {nome} \n  --> {variacao} \n  ---->{sku}") 
        custo = input("\ninforme o valor do custo do produto: ")
        sheet_produtos[f"E{linha_inserir}"] = float(custo)
        workbook_new.save("RELATORIO_VENDAS.xlsx")
        return float(custo)
        
        
    def save_pedidos(infos_pedidos, id_igual = False):
        workbook_pedidos = openpyxl.load_workbook("RELATORIO_VENDAS.xlsx")
        sheet_vendas = workbook_pedidos["VENDAS"]
        df_vendas = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="VENDAS")
        linha_inserir_vendas = sheet_vendas.max_row + 1
        # print(f"Pedido de ID: {infos_pedidos[0]} Salvando...")
        sheet_vendas[f"B{linha_inserir_vendas}"] = f"{infos_pedidos[0]}"
        sheet_vendas[f"C{linha_inserir_vendas}"] = f"{infos_pedidos[1]}"
        sheet_vendas[f"D{linha_inserir_vendas}"] = f"{infos_pedidos[2]}"
        sheet_vendas[f"E{linha_inserir_vendas}"] = f"{infos_pedidos[3]}"
        sheet_vendas[f"F{linha_inserir_vendas}"] = f"{infos_pedidos[4]}"
        sheet_vendas[f"G{linha_inserir_vendas}"] = f"{infos_pedidos[5]}"
        sheet_vendas[f"H{linha_inserir_vendas}"] = f"{infos_pedidos[6]}"
        sheet_vendas[f"I{linha_inserir_vendas}"] = f"{infos_pedidos[7]}"
        sheet_vendas[f"J{linha_inserir_vendas}"] = f"{infos_pedidos[8]}"
        sheet_vendas[f"K{linha_inserir_vendas}"] = f"{infos_pedidos[9]}"
        sheet_vendas[f"L{linha_inserir_vendas}"] = f"{infos_pedidos[10]}"
        sheet_vendas[f"M{linha_inserir_vendas}"] = f"{infos_pedidos[11]}"
        sheet_vendas[f"N{linha_inserir_vendas}"] = f"{infos_pedidos[12]}"
        sheet_vendas[f"O{linha_inserir_vendas}"] = f"{infos_pedidos[13]}"
        sheet_vendas[f"P{linha_inserir_vendas}"] = f"{infos_pedidos[14]}"
        sheet_vendas[f"Q{linha_inserir_vendas}"] = f"{infos_pedidos[15]}"
        sheet_vendas[f"R{linha_inserir_vendas}"] = f"{infos_pedidos[16]}"
                
        
        workbook_pedidos.save("RELATORIO_VENDAS.xlsx")
      

    def start_main():
        caminho_pay = "pay.xlsx"
        df_pay = pd.read_excel(caminho_pay, sheet_name="Sheet1")
        
        ## Obtenção dos ids Sacados
        start_saque = 0
        id_sacados = []
        print("Obtendo ids Sacados...")        
        for ids in df_pay.itertuples():
            if ids._3 == "Saque":
                start_saque +=1
                continue
            if start_saque == 1:
                id_sacados.append((ids._2, # 0 Tipo Transação
                                   ids._3, # 1 Descrição
                                   ids._4, # 2 ID
                                   ids._5, # 3 Direção dinehiro
                                   ids._6) # 4 Valor
                                  )
        
        caminho_all_oders = "all.xlsx"
        df_all_oders = pd.read_excel(caminho_all_oders, sheet_name="orders")
        
        list_de_infos = df_all_oders["ID do pedido"].to_list()
        
        print(list_de_infos)
        tt_ids = 1
        valor_custo_total = 0
        subtotal_pedido = 0
        id_selecionado = None
              
        for n_idsacado in id_sacados:
            
            if n_idsacado[2] not in list_de_infos:
                Main.save_pedidos([
                    f"{n_idsacado[2]}", # 0 ID Pedido
                    f"Não Encontrado", # 1 Status PEdido
                    f"-", # 2 Estado da Venda
                    f"-", # 3 Data de envio
                    f"-", # 4 Data Confirmado
                    f"-", # 5 Nome Produto
                    f"-", # 6 SKU Produto
                    f"-", # 7 Variação
                    f"-", # 8 Preço Original
                    f"-", # 9 Preço de Venda
                    f"-", # 10 Quantidade
                    f"-", # 11 Subtotal
                    f"-", # 12 Valor a Receber
                    round(float(n_idsacado[4]), 2), # 13 Valor Recebido
                    f"-", # 14 Custo
                    f"-", # 15 Lucro
                    f"-", # 16 5% Lucro
                ])
            else:
                os.system("cls")
                print("Obtendo iformações do pedido e salvando na tabela!! Não Feche! ......")
                for inf_obtd in df_all_oders.itertuples():
                
                    if n_idsacado[2] == inf_obtd._1:
                        print(f"OK! -> {n_idsacado[2]}")
                    
                        if str(inf_obtd._15) == "nan":
                                variacao = "S/V"
                        else:
                            variacao = str(inf_obtd._15)
                            
                            
                        if float(n_idsacado[4]) < 0:
                            pass
                        
                        if n_idsacado[0] == "Saldo da Carteira" and str(n_idsacado[1]).startswith("Renda do pedido") and str(n_idsacado[3]) == "Entrada":
                            
                            custo_produto = round(float(Main.consulta_preco([inf_obtd._13, variacao, inf_obtd._14]))*int(inf_obtd.Quantidade), 2)
                            subtotal_pedido = inf_obtd._20
                            valor_recebido = float(inf_obtd._20)- (
                                    inf_obtd._28 + inf_obtd._33 + inf_obtd._39 + inf_obtd._40 + inf_obtd._41 + inf_obtd._42)
                            # print(df_all_oders.columns)
                            if n_idsacado == id_selecionado:
                                """ --- produto igual! ---"""
                                Main.save_pedidos([
                                    f"-", # 0 ID Pedido
                                    f"-", # 1 Status PEdido
                                    f"-", # 2 Estado da Venda
                                    f"-", # 3 Data de envio
                                    f"-", # 4 Data Confirmado
                                    f"{inf_obtd._13}", # 5 Nome Produto
                                    f"{inf_obtd._14}", # 6 SKU Produto
                                    f"{variacao}", # 7 Variação
                                    round(float(inf_obtd._16), 2), # 8 Preço Original
                                    round(float(inf_obtd._17), 2), # 9 Preço de Venda
                                    int(inf_obtd.Quantidade), # 10 Quantidade
                                    "-", # 11 Subtotal
                                    "-", # 12 valor a receber
                                    "-", # 13 Valor Recebido
                                    "-", # 14 Custo
                                    "-", # 15 Lucro
                                    "-", # 16 5% Lucro
                                ])

                            else:
                                tt_igual = (df_all_oders['ID do pedido'].to_list()).count(F"{n_idsacado[2]}")
                                # print(tt_igual)
                                
                                # os.system("Pause")
                                if tt_igual > 1:
                                    # print("Entrou do if")
                                    valor_custo_total = 0     
                                    subtotal_pedido = 0
                                    valor_recebido_total = 0  
                                    # valor_recebido = 0   
                                    # lt = df_all_oders.to_dict()
                                    # print(lt)                      
                                    for pas in df_all_oders.itertuples():
                                        if str(pas._1) == f"{n_idsacado[2]}":
                                            variacao1 = pas._15
                                            
                                            if str(variacao1) == "nan":
                                                variacao1 = "S/V"
                                            else:
                                                variacao1 = str(pas._15)
                                            # for passe in pas.coll:
                                            #     print(passe)
                                                
                                            print(pas._20)
                                            
                                            valor_custo_total += float(Main.consulta_preco([pas._13, variacao1, pas._14])*int(pas.Quantidade))
                                            subtotal_pedido += float(pas._20)
                                            # valor_recebido_un =  (pas._20 - (
                                            #     float(pas._28) + float(pas._33) + float(pas._39) + float(pas._40) + float(pas._41) + float(pas._42)))  
                                            # valor_recebido_total += valor_recebido_un
                                            gastos_pago = float(pas._28) + float(pas._33) + float(pas._39) + float(pas._40) + float(pas._41) + float(pas._42)
                                            
                                    custo_produto = round(float(valor_custo_total), 2)
                                    
                                    valor_recebido = subtotal_pedido - gastos_pago
                                    print(valor_recebido)
                                    print(gastos_pago)
                                    
                                    
                                    # os.system("Pause")
                                                                            
                                
                                per_lucro = (custo_produto*100)/valor_recebido
                                lucro = (valor_recebido - custo_produto)
                                
                                Main.save_pedidos([
                                    f"{inf_obtd._1}", # 0 ID Pedido
                                    f"{inf_obtd._2}", # 1 Status PEdido
                                    f"{inf_obtd.UF}", # 2 Estado da Venda
                                    f"{inf_obtd._9}", # 3 Data de envio
                                    f"{inf_obtd._57}", # 4 Data Confirmado
                                    f"{inf_obtd._13}", # 5 Nome Produto
                                    f"{inf_obtd._14}", # 6 SKU Produto
                                    f"{variacao}", # 7 Variação
                                    round(float(inf_obtd._16), 2), # 8 Preço Original
                                    round(float(inf_obtd._17), 2), # 9 Preço de Venda
                                    int(inf_obtd.Quantidade), # 10 Quantidade
                                    round(float(subtotal_pedido), 2), # 11 Subtotal
                                    round(float(valor_recebido), 2), # 12 Valor a Receber
                                    round(float(n_idsacado[4]), 2), # 13 Valor Recebido
                                    round(float(custo_produto), 2), # 14 Custo
                                    round(float(lucro), 2), # 15 Lucro
                                    round(float(per_lucro), 2), # 16 5% Lucro
                                ])
                            
                        else:
                            Main.save_pedidos([
                                    f"{n_idsacado[2]}", # 0 ID Pedido
                                    f"Outros", # 1 Status PEdido
                                    f"{inf_obtd.UF}", # 2 Estado da Venda
                                    f"{inf_obtd._9}", # 3 Data de envio
                                    f"{inf_obtd._57}", # 4 Data Confirmado
                                    f"-", # 5 Nome Produto
                                    f"-", # 6 SKU Produto
                                    f"-", # 7 Variação
                                    f"-", # 8 Preço Original
                                    f"-", # 9 Preço de Venda
                                    f"-", # 10 Quantidade
                                    f"-", # 11 Subtotal
                                    f"-", # 12 Valor a Receber
                                    round(float(n_idsacado[4]), 2), # 13 Valor recebido
                                    f"-", # 14 Custo
                                    f"-", # 15 Lucro
                                    f"-", # 16 5% Lucro
                                ])
  
                        
                            print("ID: ", n_idsacado[2], "Salvo!!")
                        
                        tt_ids +=1
                        id_selecionado = n_idsacado
        
        print("Finalizado com Sucesso!!")                  
                
Main.start_main()
