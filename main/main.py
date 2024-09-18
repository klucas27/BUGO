import openpyxl.workbook
import pandas as pd
import openpyxl
print("Software para analises de dados!\n")


class Bugo():
    def __init__(self) -> None:
        escolha = input("Informe a opção: ")
        if escolha == 1:
            self.get_value_produtos
        pass
    
    def get_value_produtos():
        
        # caminho = input("Informe o caminho: ")
        caminho = "C:\\Users\\PC FEIRAO 03\\Desktop\\Order.shipping.20240912_20240912.xlsx"
        workbook = openpyxl.load_workbook(caminho)
        pnl = workbook['orders']
        # df = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="PRUDUTOS")
        workbook2 = openpyxl.load_workbook("RELATORIO_VENDAS.xlsx")
        sheet_produtos = workbook2['PRUDUTOS']
                
        # print(pnl.max_row)
        ctt = 2
        ctt2 = 4
        while ctt <= pnl.max_row-80:
            info_pedido = {
                f"{pnl[f'A{ctt}'].value}": [
                    # Sobre o Pedido  
                    pnl[f'A{ctt}'].value,  # ID 0
                    pnl[f'J{ctt}'].value,  # Data Pagamento 1
                    
                    # Produtos 
                    pnl[f'L{ctt}'].value,  # Nome do Produto 2
                    pnl[f'M{ctt}'].value,  # SKU do Produto 3
                    pnl[f'N{ctt}'].value,  # Variacao Produto 4
                    pnl[f'O{ctt}'].value,  # Preco Original Produto 5
                    pnl[f'P{ctt}'].value,  # Preco Acordadeo do Produto 6
                    pnl[f'Q{ctt}'].value,  # Quantidade do Produto 7
                    pnl[f'AL{ctt}'].value,  #Taxa Envio Reverso 8
                    pnl[f'AM{ctt}'].value,  # Taxa Transacao 9
                    pnl[f'AN{ctt}'].value,  # Taxa Comissao 10
                    pnl[f'AO{ctt}'].value,  #  Taxa Servico 11
                    
                    # Sobre o Cliente
                    pnl[f'AR{ctt}'].value,  # Username
                    pnl[f'AY{ctt}'].value,  # Cidade Cliente
                    pnl[f'AZ{ctt}'].value,  # Estado Cliente
                ]
            }
            
            for pas in info_pedido.values():
                df = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="PRUDUTOS")
                if pas[3] in df.iloc[:, 1].values:
                    pass
                else:
                    sheet_produtos[f"C{ctt2}"] = pas[2]
                    if pas[3] == "":
                        valor_sku = input(f"Insira o sku de {pas[2]}")
                        sheet_produtos[f"B{ctt2}"] = valor_sku
                    else:
                        sheet_produtos[f"B{ctt2}"] = pas[3]
                        
                    custo_produto = input(f"Insira o Valor do produto {pas[2]} de SKU: {pas[3]}: \n")
                    sheet_produtos[f"D{ctt2}"] = custo_produto
                    workbook2.save("RELATORIO_VENDAS.xlsx")
                    ctt2 += 1
                
                # sheet[f"C{ctt2}"] = pas[1]
                # sheet[f"C{ctt2}"] = pas[1]
                print(pas[0])
            ctt += 1
            
            
        workbook2.save("RELATORIO_VENDAS.xlsx")
    
    def pass_vendas_in_table():
        
        
        
        
if __name__ == "__main__":
    Bugo.get_value_produtos()
    