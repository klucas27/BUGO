import openpyxl.workbook
import pandas as pd
import openpyxl


class Main2:
    def __init__(self) -> None:
        pass
    
    def save_produtos(produto, id_pedido):
        workbook_new = openpyxl.load_workbook("RELATORIO_VENDAS.xlsx")
        sheet_produtos = workbook_new["PRUDUTOS"]
        sheet_vendas = workbook_new["VENDAS"]
        df = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="PRUDUTOS")
        linha_inserir = sheet_produtos.max_row + 1
        linha_inserir_vendas = sheet_vendas.max_row + 1    
        
        #print(sheet_produtos.max_row)
        if produto[0] in df.iloc[:, 2].values:
            pass
        else:
            sheet_produtos[f"C{linha_inserir}"] = f"{produto[0]}"
            sheet_produtos[f"B{linha_inserir}"] = f"{produto[1]}"
            custo = input("informe o valor do custo do produto: ")
            sheet_produtos[f"D{linha_inserir}"] = custo
        workbook_new.save("RELATORIO_VENDAS.xlsx")
        
        
        
        sheet_vendas[f"B{linha_inserir_vendas}"] = f"{id_pedido}"  # ID
        sheet_vendas[f"C{linha_inserir_vendas}"] = f"{produto[1]}"  # SKU
        sheet_vendas[f"D{linha_inserir_vendas}"] = f"{produto[0]}"  #  Nome
        sheet_vendas[f"E{linha_inserir_vendas}"] = f"{produto[2]}"  # Data PEdido
        sheet_vendas[f"F{linha_inserir_vendas}"] = f"{produto[3]}"  # Valor Pedido
        
        valor_pago_plataforma = float(produto[7]) + float(produto[8]) + float(produto[9]) + float(produto[10]) + float(produto[6])
        sheet_vendas[f"G{linha_inserir_vendas}"] = f"{valor_pago_plataforma}"  #Valor Pago Plataforma
        custo_produto = 0
        line_find = 0
        #print(sheet_vendas["C"][3].value)
        for linha in sheet_produtos['C']:
            # print(linha.value)
            if linha.value == produto[0]:
                print(linha.value)
                custo_produto = sheet_produtos[f"D{line_find+1}"].value
            line_find +=1
            
        print(custo_produto)
        sheet_vendas[f"H{linha_inserir_vendas}"] = f"{custo_produto}"  # Custo Produto
       # sheet_vendas[f"I{linha_inserir_vendas}"] = f"{produto[0]}"  #Lucro final
        workbook_new.save("RELATORIO_VENDAS.xlsx")
    
    def save_pedidos(linha_inserir, infos_pedidos):
        pass
        
        
    def obtendo_valores():     
        
        caminho = "pay07-14.xlsx"
        caminho_all = "all19-08a18-09.xlsx"
        workbook = openpyxl.load_workbook(caminho)
        workbook_all = openpyxl.load_workbook(caminho_all)
    
        pnl = workbook['Sheet1']
        pnl_all = workbook_all['orders']
        
        ctt = 0
        cont = 19
        total_valor = 0
        print("start")
        while ctt != 2:                           
            if f"{pnl[f"C{cont}"].value}" != "Saque" and ctt == 1:
                ### Valores que não sao o saque
                # print(pnl[f"D{cont}"].value, pnl[f"F{cont}"].value)
                total_valor += float(pnl[f"F{cont}"].value)
                
                linha_achada = 1
                for linha in pnl_all["A"]:
                    if linha.value == pnl[f"D{cont}"].value:
                        print(pnl_all[f"M{linha_achada}"].value)
                        Main2.save_produtos([pnl_all[f"M{linha_achada}"].value,  #  0 Nome Produto
                                            pnl_all[f"N{linha_achada}"].value,  #  1 SKU Produto
                                            pnl_all[f"K{linha_achada}"].value,  #  2 data e hora pagamento (pedido caiu para embalar)
                                            pnl_all[f"Q{linha_achada}"].value,  #  3 Peço de venda
                                            pnl_all[f"R{linha_achada}"].value,  #  4 Quantidade
                                            pnl_all[f"T{linha_achada}"].value,  #  5 SubTotal  
                                            pnl_all[f"AB{linha_achada}"].value,  #  6 Cupom Vendedor  
                                            pnl_all[f"AC{linha_achada}"].value,  #  7 Coin Cashback  
                                            pnl_all[f"AF{linha_achada}"].value,  #  8 Desconto + por -  
                                            pnl_all[f"AO{linha_achada}"].value,  #  9 Taxa comissao  
                                            pnl_all[f"AP{linha_achada}"].value,  #  10 Taxa Serviço  
                                            ], pnl_all[f"A{linha_achada}"].value)
                    linha_achada += 1
                
                  
            
                
            elif f"{pnl[f"C{cont}"].value}" == "Saque" and ctt == 0:
                total_valor -= float(pnl[f"H{cont}"].value)   
                
            if f"{pnl[f"C{cont}"].value}" == "Saque":
                ctt+=1
            cont+=1
        # print(pnl["A"].value)
        print(f"TOTAL DE SAQUE: {float(total_valor):.2f}")
Main2.obtendo_valores()