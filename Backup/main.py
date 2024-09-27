import openpyxl.workbook
import pandas as pd
import openpyxl
import os

# 240907C70PXA8X

class Main:
    def __init__(self) -> None:
        pass
   
    def consulta_preco(produto):
        """Produto = [
            0 - nome
            1 - variacao
            2 - sku
        ]

        """
        workbook_new = openpyxl.load_workbook("RELATORIO_VENDAS.xlsx")
        sheet_produtos = workbook_new["PRODUTOS"]
        df = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="PRODUTOS")
        
        linha = 0
        valor = 0
        
        for produtos in sheet_produtos['C']:
            # print(produto[0])
            # print(produto[1])
            if produto[0] == produtos.value and produto[1] == sheet_produtos['D'][linha].value:
                valor = sheet_produtos['E'][linha].value
                return float(valor)
            
            linha += 1        

        valor = Main.save_produtos([produto[2], # sku
                                    produto[0],  # nome 
                                    produto[1]]) # variacao
        return float(valor)
        
    
    def save_produtos(produto):
        """SALVA PRODUTOS NA TABELA RELATORIO VENDAS

        Args:
            produto (_list_): Eecebe lista com informações do produto
        """
        workbook_new = openpyxl.load_workbook("RELATORIO_VENDAS.xlsx")
        sheet_produtos = workbook_new["PRODUTOS"]
        df = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="PRODUTOS")
        nome = produto[1]
        variacao = produto[2]
        sku = produto[0]
        custo = 0
        
        linha_inserir = sheet_produtos.max_row + 1

        if variacao == "" or variacao == "nan":
            variacao = "S/V"
            
        # if nome in df.iloc[:, 2].values and variacao in df.iloc[:, 3].values:
        #     print("Entrou na 1 condi")
        # elif nome in df.iloc[:, 2].values and variacao == "S/V":
        #     print("Entrou na 2 condi")
        # else:
        sheet_produtos[f"B{linha_inserir}"] = f"{sku}"   #  SKU
        sheet_produtos[f"C{linha_inserir}"] = f"{nome}"   # Nome
        sheet_produtos[f"D{linha_inserir}"] = f"{variacao}"   # variação
        os.system("cls")
        print(f"\n -> {nome} \n  --> {variacao} \n  ---->{sku}") 
        custo = input("\ninforme o valor do custo do produto: ")
        sheet_produtos[f"E{linha_inserir}"] = float(custo)
        workbook_new.save("RELATORIO_VENDAS.xlsx")
        return float(custo)
        
        
    def save_pedidos(infos_pedidos, id_igual = False):
        workbook_pedidos = openpyxl.load_workbook("RELATORIO_VENDAS.xlsx")
        sheet_vendas = workbook_pedidos["VENDAS"]
        df_vendas = pd.read_excel("RELATORIO_VENDAS.xlsx", sheet_name="VENDAS")
        linha_inserir_vendas = sheet_vendas.max_row + 1
        # print(f"Pedido de ID: {infos_pedidos[0]} Salvando...")
        sheet_vendas[f"B{linha_inserir_vendas}"] = f"{infos_pedidos[0]}"
        sheet_vendas[f"C{linha_inserir_vendas}"] = f"{infos_pedidos[1]}"
        sheet_vendas[f"D{linha_inserir_vendas}"] = f"{infos_pedidos[2]}"
        sheet_vendas[f"E{linha_inserir_vendas}"] = f"{infos_pedidos[3]}"
        sheet_vendas[f"F{linha_inserir_vendas}"] = f"{infos_pedidos[4]}"
        sheet_vendas[f"G{linha_inserir_vendas}"] = f"{infos_pedidos[5]}"
        sheet_vendas[f"H{linha_inserir_vendas}"] = f"{infos_pedidos[6]}"
        sheet_vendas[f"I{linha_inserir_vendas}"] = f"{infos_pedidos[7]}"
        sheet_vendas[f"J{linha_inserir_vendas}"] = f"{infos_pedidos[8]}"
        sheet_vendas[f"K{linha_inserir_vendas}"] = f"{infos_pedidos[9]}"
        sheet_vendas[f"L{linha_inserir_vendas}"] = f"{infos_pedidos[10]}"
        sheet_vendas[f"M{linha_inserir_vendas}"] = f"{infos_pedidos[11]}"
        sheet_vendas[f"N{linha_inserir_vendas}"] = f"{infos_pedidos[12]}"
        sheet_vendas[f"O{linha_inserir_vendas}"] = f"{infos_pedidos[13]}"
        sheet_vendas[f"P{linha_inserir_vendas}"] = f"{infos_pedidos[14]}"
        sheet_vendas[f"Q{linha_inserir_vendas}"] = f"{infos_pedidos[15]}"
        sheet_vendas[f"R{linha_inserir_vendas}"] = f"{infos_pedidos[16]}"
        
        workbook_pedidos.save("RELATORIO_VENDAS.xlsx")
      

    def start_main():
        caminho_pay = "pay.xlsx"
        df_pay = pd.read_excel(caminho_pay, sheet_name="Sheet1")
        
        ## Obtenção dos ids Sacados
        start_saque = 0
        id_sacados = []
        print("Obtendo ids Sacados...")        
        for ids in df_pay.itertuples():
            # id_sacados.append((ids._4, ids._6))
            if ids._3 == "Saque":
                start_saque +=1
                continue
            if start_saque == 1:
                id_sacados.append((ids._2, # 0 Tipo Transação
                                   ids._3, # 1 Descrição
                                   ids._4, # 2 ID
                                   ids._5, # 3 Direção dinehiro
                                   ids._6) # 4 Valor
                                  )
        
        ## Obtenção das informações dos pedidos conforme por id
        caminho_all_oders = "all.xlsx"
        df_all_oders = pd.read_excel(caminho_all_oders, sheet_name="orders")
        
        list_de_infos = df_all_oders["ID do pedido"].to_list()
        
        print(list_de_infos)
        tt_ids = 1
        valor_custo_total = 0
        subtotal_pedido = 0
        id_selecionado = None
              
        for n_idsacado in id_sacados:
            
            if n_idsacado[2] not in list_de_infos:
                Main.save_pedidos([
                    f"{n_idsacado[2]}", # 0 ID Pedido
                    f"Não Encontrado", # 1 Status PEdido
                    f"-", # 2 Estado da Venda
                    f"-", # 3 Data de envio
                    f"-", # 4 Data Confirmado
                    f"-", # 5 Nome Produto
                    f"-", # 6 SKU Produto
                    f"-", # 7 Variação
                    f"-", # 8 Preço Original
                    f"-", # 9 Preço de Venda
                    f"-", # 10 Quantidade
                    f"-", # 11 Subtotal
                    f"-", # 12 Valor a Receber
                    round(float(n_idsacado[4]), 2), # 13 Valor Recebido
                    f"-", # 14 Custo
                    f"-", # 15 Lucro
                    f"-", # 16 5% Lucro
                ])
            else:
                os.system("cls")
                print("Obtendo iformações do pedido e salvando na tabela!! Não Feche! ......")
                for inf_obtd in df_all_oders.itertuples():
                
                    if n_idsacado[2] == inf_obtd._1:
                        print(f"OK! -> {n_idsacado[2]}")
                    
                        if str(inf_obtd._15) == "nan":
                                variacao = "S/V"
                        else:
                            variacao = str(inf_obtd._15)
                            
                            
                        if float(n_idsacado[4]) < 0:
                            pass
                        
                        if n_idsacado[0] == "Saldo da Carteira" and str(n_idsacado[1]).startswith("Renda do pedido") and str(n_idsacado[3]) == "Entrada":
                            
                            custo_produto = round(float(Main.consulta_preco([inf_obtd._13, variacao, inf_obtd._14]))*int(inf_obtd.Quantidade), 2)
                            subtotal_pedido = inf_obtd._20
                            valor_recebido = float(inf_obtd._20)- (
                                    inf_obtd._28 + inf_obtd._33 + inf_obtd._39 + inf_obtd._40 + inf_obtd._41 + inf_obtd._42)
                            # print(df_all_oders.columns)
                            if n_idsacado == id_selecionado:
                                """ --- produto igual! ---"""
                                Main.save_pedidos([
                                    f"-", # 0 ID Pedido
                                    f"-", # 1 Status PEdido
                                    f"-", # 2 Estado da Venda
                                    f"-", # 3 Data de envio
                                    f"-", # 4 Data Confirmado
                                    f"{inf_obtd._13}", # 5 Nome Produto
                                    f"{inf_obtd._14}", # 6 SKU Produto
                                    f"{variacao}", # 7 Variação
                                    round(float(inf_obtd._16), 2), # 8 Preço Original
                                    round(float(inf_obtd._17), 2), # 9 Preço de Venda
                                    int(inf_obtd.Quantidade), # 10 Quantidade
                                    "-", # 11 Subtotal
                                    "-", # 12 valor a receber
                                    "-", # 13 Valor Recebido
                                    "-", # 14 Custo
                                    "-", # 15 Lucro
                                    "-", # 16 5% Lucro
                                ])

                            else:
                                tt_igual = (df_all_oders['ID do pedido'].to_list()).count(F"{n_idsacado[2]}")
                                # print(tt_igual)
                                
                                # os.system("Pause")
                                if tt_igual > 1:
                                    # print("Entrou do if")
                                    valor_custo_total = 0     
                                    subtotal_pedido = 0
                                    valor_recebido_total = 0  
                                    # valor_recebido = 0   
                                    # lt = df_all_oders.to_dict()
                                    # print(lt)                      
                                    for pas in df_all_oders.itertuples():
                                        if str(pas._1) == f"{n_idsacado[2]}":
                                            variacao1 = pas._15
                                            
                                            if str(variacao1) == "nan":
                                                variacao1 = "S/V"
                                            else:
                                                variacao1 = str(pas._15)
                                            # for passe in pas.coll:
                                            #     print(passe)
                                                
                                            print(pas._20)
                                            
                                            valor_custo_total += float(Main.consulta_preco([pas._13, variacao1, pas._14])*int(pas.Quantidade))
                                            subtotal_pedido += float(pas._20)
                                            # valor_recebido_un =  (pas._20 - (
                                            #     float(pas._28) + float(pas._33) + float(pas._39) + float(pas._40) + float(pas._41) + float(pas._42)))  
                                            # valor_recebido_total += valor_recebido_un
                                            gastos_pago = float(pas._28) + float(pas._33) + float(pas._39) + float(pas._40) + float(pas._41) + float(pas._42)
                                            
                                    custo_produto = round(float(valor_custo_total), 2)
                                    
                                    valor_recebido = subtotal_pedido - gastos_pago
                                    print(valor_recebido)
                                    print(gastos_pago)
                                    
                                    
                                    # os.system("Pause")
                                                                            
                                
                                per_lucro = (custo_produto*100)/valor_recebido
                                lucro = (valor_recebido - custo_produto)
                                
                                Main.save_pedidos([
                                    f"{inf_obtd._1}", # 0 ID Pedido
                                    f"{inf_obtd._2}", # 1 Status PEdido
                                    f"{inf_obtd.UF}", # 2 Estado da Venda
                                    f"{inf_obtd._9}", # 3 Data de envio
                                    f"{inf_obtd._57}", # 4 Data Confirmado
                                    f"{inf_obtd._13}", # 5 Nome Produto
                                    f"{inf_obtd._14}", # 6 SKU Produto
                                    f"{variacao}", # 7 Variação
                                    round(float(inf_obtd._16), 2), # 8 Preço Original
                                    round(float(inf_obtd._17), 2), # 9 Preço de Venda
                                    int(inf_obtd.Quantidade), # 10 Quantidade
                                    round(float(subtotal_pedido), 2), # 11 Subtotal
                                    round(float(valor_recebido), 2), # 12 Valor a Receber
                                    round(float(n_idsacado[4]), 2), # 13 Valor Recebido
                                    round(float(custo_produto), 2), # 14 Custo
                                    round(float(lucro), 2), # 15 Lucro
                                    round(float(per_lucro), 2), # 16 5% Lucro
                                ])
                            
                        else:
                            Main.save_pedidos([
                                    f"{n_idsacado[2]}", # 0 ID Pedido
                                    f"Outros", # 1 Status PEdido
                                    f"{inf_obtd.UF}", # 2 Estado da Venda
                                    f"{inf_obtd._9}", # 3 Data de envio
                                    f"{inf_obtd._57}", # 4 Data Confirmado
                                    f"-", # 5 Nome Produto
                                    f"-", # 6 SKU Produto
                                    f"-", # 7 Variação
                                    f"-", # 8 Preço Original
                                    f"-", # 9 Preço de Venda
                                    f"-", # 10 Quantidade
                                    f"-", # 11 Subtotal
                                    f"-", # 12 Valor a Receber
                                    round(float(n_idsacado[4]), 2), # 13 Valor recebido
                                    f"-", # 14 Custo
                                    f"-", # 15 Lucro
                                    f"-", # 16 5% Lucro
                                ])
                                                        
                            
                            
                            
                            
                        #     Main.save_pedidos([
                        #     new_infos[ctt][0], #0
                        #     new_infos[ctt][1], #1
                        #     new_infos[ctt][18],  #2
                        #     new_infos[ctt][2],  # 3
                        #     new_infos[ctt][3],  # 4
                        #     new_infos[ctt][4],  # 5
                        #     new_infos[ctt][5],  # 6
                        #     new_infos[ctt][6],  # 7
                        #     new_infos[ctt][7],  # 8
                        #     round(float(new_infos[ctt][8]), 2),  #9
                        #     new_infos[ctt][9],  # 10
                        #     round(float(total_venda), 2),  # 11
                        #     round(float(valor_recebido), 2),  #12
                        #     round(float(custo_total), 2),  #13
                        #     round(float(lucro_final), 2),  # 14
                        #     round(float(porct_lucro), 2),  # 15
                                                
                        # ])
        
                        
                            print("ID: ", n_idsacado[2], "Salvo!!")
                        
                        tt_ids +=1
                        id_selecionado = n_idsacado
                    
                # else:
                #     print(f"NÃO OK -> {n_idsacado[2]}")
                #     Main.save_pedidos([
                #                 f"{n_idsacado[2]}", # 0 ID Pedido
                #                 f"Não Encontrado", # 1 Status PEdido
                #                 f"-", # 2 Estado da Venda
                #                 f"-", # 3 Data de envio
                #                 f"-", # 4 Data Confirmado
                #                 f"-", # 5 Nome Produto
                #                 f"-", # 6 SKU Produto
                #                 f"-", # 7 Variação
                #                 f"-", # 8 Preço Original
                #                 f"-", # 9 Preço de Venda
                #                 f"-", # 10 Quantidade
                #                 f"-", # 11 Subtotal
                #                 round(float(n_idsacado[4]), 2), # 12 Valor Recebido
                #                 f"-", # 13 Custo
                #                 f"-", # 14 Lucro
                #                 f"-", # 5% Lucro
                #             ])
                #     break
                    
        # os.system("cls")
        print("Finalizado com Sucesso!")
        # print(tt_ids)
        
        
        # for get_infos in df_all_oders.itertuples():
            # for ver_ids_sacados in id_sacados:
                # print(ver_ids_sacados[0])
        
        """
        ## Obtenção das informações dos pedidos conforme por id
        print("Obtendo informações do pedido...")
        caminho_all_oders = "all.xlsx"
        
        df_all_oders = pd.read_excel(caminho_all_oders, sheet_name="orders")
        # print(id_sacados.count("240831QTPQHE7X"))
        infos_obtidas = []
        id_pedido_only = []
        for pedidos_all in df_all_oders.itertuples():
            if id_sacados.count(f"{pedidos_all._1}") > 0:
                # print(pedidos_all)
                id_pedido_only.append(str(pedidos_all._1))
                variacao = None
                if str(pedidos_all._15) == "nan":
                    variacao = "S/V"
                else:
                    variacao = str(pedidos_all._15)
                    
                custo_produto = round(float(Main.consulta_preco([pedidos_all._13, variacao, pedidos_all._14])), 2)
                infos_obtidas.append([
                    pedidos_all._1,  #  0 ID
                    pedidos_all._2,  #  1 Status
                    pedidos_all._9,  #  2 dia do envio
                    pedidos_all._11,  # 3 dia pedido confirmado
                    pedidos_all._13,  #  4 Nome do produto
                    pedidos_all._14,  #  5 sku
                    variacao,  #  6 variação
                    pedidos_all._16,  #  7 preço original
                    pedidos_all._17,  #  8 preço de venda
                    pedidos_all.Quantidade,  #  9 quantidade
                    pedidos_all._20,  #  10 Sbtotal
                    pedidos_all._28,  #  11 gasto cupom
                    pedidos_all._33,  #  12 desconto + po -
                    pedidos_all._36,  #  13 Valor Total
                    pedidos_all._39,  #  14 Taxa reversa
                    pedidos_all._40,  #  15 Taxa transação
                    pedidos_all._41,  # 16 Taxa comissao
                    pedidos_all._42,  #  17 Taxa de serviço
                    pedidos_all.UF,  #  18 Estado do cliente
                    pedidos_all._57,  #  19 data completado
                    float(custo_produto)*int(pedidos_all.Quantidade),  # 20 Custo produto *                    
                    ])
    
        ## obtendo valores de custo
        print("Ordenando os pedidos...")
        caminho_save_relatorio = "RELATORIO_VENDAS.xlsx"
        df_produtos_save = pd.read_excel(caminho_save_relatorio, sheet_name="PRODUTOS")
        new_infos = []
        for pedidos_ordenados in id_sacados:
            for pedidos_desordenados in infos_obtidas:
                #print(pedidos_ordenados)
                if str(pedidos_desordenados[0]) == str(pedidos_ordenados):
                    new_infos.append(pedidos_desordenados)
                
                
        print("Salvando os pedidos na tabela!(Pode Demorar Bastante)...")
        ctt = 0
        while ctt < len(new_infos):
            print("ID: ", str(new_infos[ctt][0]), " Salvando...")
            tt_prod = int(id_pedido_only.count(str(new_infos[ctt][0])))

            total_venda = 0
            custo_total = 0
            while tt_prod >= 1:
                total_venda += new_infos[ctt][10]
                custo_total += new_infos[ctt][20]
                
                if tt_prod == 1:
                    valor_recebido = total_venda - (
                        float(new_infos[ctt][14]) +
                        float(new_infos[ctt][15]) +
                        float(new_infos[ctt][16]) +
                        float(new_infos[ctt][11]) +
                        float(new_infos[ctt][17])
                    )
                    
                    lucro_final = valor_recebido - custo_total
                    porct_lucro = lucro_final*100/custo_total
                    
                    Main.save_pedidos([
                        new_infos[ctt][0], #0
                        new_infos[ctt][1], #1
                        new_infos[ctt][18],  #2
                        new_infos[ctt][2],  # 3
                        new_infos[ctt][3],  # 4
                        new_infos[ctt][4],  # 5
                        new_infos[ctt][5],  # 6
                        new_infos[ctt][6],  # 7
                        new_infos[ctt][7],  # 8
                        round(float(new_infos[ctt][8]), 2),  #9
                        new_infos[ctt][9],  # 10
                        round(float(total_venda), 2),  # 11
                        round(float(valor_recebido), 2),  #12
                        round(float(custo_total), 2),  #13
                        round(float(lucro_final), 2),  # 14
                        round(float(porct_lucro), 2),  # 15
                                               
                    ])
                else:
                    Main.save_pedidos([
                        "||", #0
                        "||", #1
                        new_infos[ctt][18],  #2
                        new_infos[ctt][2],  # 3
                        new_infos[ctt][3],  # 4
                        new_infos[ctt][4],  # 5
                        new_infos[ctt][5],  # 6
                        new_infos[ctt][6],  # 7
                        new_infos[ctt][7],  # 8
                        round(float(new_infos[ctt][8]), 2),  #9
                        new_infos[ctt][9],  # 10
                        "||",  # 11
                        "||",  #12
                        "||",  #13
                        "||",  # 14
                        "||",  # 15
                    ])  
                    ctt+=1
                tt_prod -= 1
            ctt += 1
        """
                
Main.start_main()
